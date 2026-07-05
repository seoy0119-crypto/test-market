# link_list_template.py
#
# Reference snippet for Step 2b (link-document delivery) in SKILL.md. Adapt `data` to what you
# actually collected — this is here so the openpyxl boilerplate (header styling, hyperlink
# cells, column widths) doesn't need to be reinvented each time.
#
# `data` shape: {"<키워드 label>": [{"pin": "<pin url>", "img": "<image url>"}, ...], ...}

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def build_link_doc(data: dict, out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "무드보드 이미지 목록"

    headers = ["No", "키워드", "핀 링크", "이미지 URL"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="4A4A4A")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row, no = 2, 1
    for label, items in data.items():
        for item in items:
            ws.cell(row=row, column=1, value=no).font = Font(name="Arial")
            ws.cell(row=row, column=2, value=label).font = Font(name="Arial")
            c3 = ws.cell(row=row, column=3, value=item["pin"])
            c3.hyperlink = item["pin"]
            c3.font = Font(name="Arial", color="0563C1", underline="single")
            c4 = ws.cell(row=row, column=4, value=item["img"])
            c4.hyperlink = item["img"]
            c4.font = Font(name="Arial", color="0563C1", underline="single")
            row += 1
            no += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 60
    ws.freeze_panes = "A2"
    wb.save(out_path)


if __name__ == "__main__":
    # example usage
    example = json.loads(open("data.json").read())
    build_link_doc(example, "moodboard_links.xlsx")
